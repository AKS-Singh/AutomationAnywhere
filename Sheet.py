import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def execute_excel(file_name):
    wb = xl.load_workbook(file_name)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        mrp = sheet.cell(row, 3)
        discount = sheet.cell(row, 4)
        sale_price = mrp.value - discount.value
        sale_price_cell = sheet.cell(row, 5)
        sale_price_cell.value = sale_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=5, max_col=5)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'f2')

    wb.save(file_name)


file = input("Enter the file name you want to process ")
execute_excel(file)
